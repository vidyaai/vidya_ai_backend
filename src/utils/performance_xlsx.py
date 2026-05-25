"""XLSX workbook builder for class performance reports with embedded native charts."""
from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.performance_stats import (
    compute_histogram,
    compute_weighted_totals,
)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="0F766E")  # teal-700
_CENTER = Alignment(horizontal="center", vertical="center")
_BIN_SIZE = 10


def _autosize(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


def _style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _write_raw_data(ws, students: list[dict], assignments: list[dict], weighted_totals: list[float]) -> None:
    ws.append(["Student", "Email"] + [a["title"] for a in assignments] + ["Weighted Total"])
    _style_header_row(ws, 1, 2 + len(assignments) + 1)
    for student, total in zip(students, weighted_totals):
        scores = student.get("scores") or {}
        row = [student.get("name") or "", student.get("email") or ""]
        for a in assignments:
            v = scores.get(a["id"])
            row.append(v if v is not None else "")
        row.append(total)
        ws.append(row)
    _autosize(ws)


def _write_statistics(ws, assignments: list[dict], weightages: dict[str, float]) -> int:
    ws.append([
        "Assignment", "Weight", "Count", "Mean", "Median", "Stdev",
        "Q1", "Q3", "Min", "Max", "Submission Rate %", "On-Time Rate %",
    ])
    _style_header_row(ws, 1, 12)
    for a in assignments:
        st = a.get("stats") or {}
        w_pct = round(100.0 * weightages.get(a["id"], 0.0), 2)
        ws.append([
            a.get("title") or "",
            w_pct,
            st.get("count") or 0,
            st.get("mean"),
            st.get("median"),
            st.get("stdev"),
            st.get("q1"),
            st.get("q3"),
            st.get("min"),
            st.get("max"),
            a.get("submission_rate"),
            a.get("on_time_rate"),
        ])
    last_row = 1 + len(assignments)
    _autosize(ws)
    return last_row


def _write_histograms_sheet(ws, assignments: list[dict]) -> dict[str, tuple[int, int]]:
    """Write per-assignment histogram columns. Returns {assignment_id: (start_col, n_buckets)}."""
    bucket_labels = [b["bucket"] for b in compute_histogram([], _BIN_SIZE)]
    n = len(bucket_labels)
    ws.cell(row=1, column=1, value="Bucket").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    for i, label in enumerate(bucket_labels, start=2):
        ws.cell(row=i, column=1, value=label)
    location: dict[str, tuple[int, int]] = {}
    col = 2
    for a in assignments:
        header = ws.cell(row=1, column=col, value=a.get("title") or a["id"])
        header.font = _HEADER_FONT
        header.fill = _HEADER_FILL
        for i, bucket in enumerate(a.get("histogram") or [], start=2):
            ws.cell(row=i, column=col, value=bucket.get("count", 0))
        location[a["id"]] = (col, n)
        col += 1
    _autosize(ws)
    return location


def _add_histogram_charts(
    charts_ws,
    hist_ws,
    assignments: list[dict],
    location: dict[str, tuple[int, int]],
) -> int:
    """Anchor a BarChart per assignment in two columns; return next free row."""
    anchor_row = 2
    col_letters = ["B", "L"]  # two charts per row
    for idx, a in enumerate(assignments):
        col_letter = col_letters[idx % 2]
        if idx > 0 and idx % 2 == 0:
            anchor_row += 16
        chart = BarChart()
        chart.type = "col"
        chart.style = 11
        chart.title = f"Score distribution — {a.get('title') or a['id']}"
        chart.y_axis.title = "Students"
        chart.x_axis.title = "Score range (%)"
        chart.legend = None
        col, n = location[a["id"]]
        data = Reference(hist_ws, min_col=col, max_col=col, min_row=1, max_row=1 + n)
        cats = Reference(hist_ws, min_col=1, min_row=2, max_row=1 + n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 15
        charts_ws.add_chart(chart, f"{col_letter}{anchor_row}")
    return anchor_row + 16


def _add_submission_chart(charts_ws, stats_ws, n_assignments: int, anchor: str) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 12
    chart.grouping = "clustered"
    chart.title = "Submission rate vs On-time rate"
    chart.y_axis.title = "Rate (%)"
    chart.x_axis.title = "Assignment"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    # Submission Rate (col 11) and On-Time Rate (col 12) from Statistics sheet
    data = Reference(
        stats_ws, min_col=11, max_col=12, min_row=1, max_row=1 + n_assignments
    )
    cats = Reference(stats_ws, min_col=1, min_row=2, max_row=1 + n_assignments)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    charts_ws.add_chart(chart, anchor)


def _add_trend_chart(charts_ws, stats_ws, n_assignments: int, anchor: str) -> None:
    chart = LineChart()
    chart.title = "Class average per assignment (chronological)"
    chart.style = 13
    chart.y_axis.title = "Mean score (%)"
    chart.x_axis.title = "Assignment"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    # Mean (col 4)
    data = Reference(
        stats_ws, min_col=4, max_col=4, min_row=1, max_row=1 + n_assignments
    )
    cats = Reference(stats_ws, min_col=1, min_row=2, max_row=1 + n_assignments)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    charts_ws.add_chart(chart, anchor)


def _add_weighted_total_chart(
    charts_ws, hist_ws, weighted_totals: list[float], start_col: int, anchor: str
) -> None:
    buckets = compute_histogram(weighted_totals, _BIN_SIZE)
    n = len(buckets)
    header = hist_ws.cell(row=1, column=start_col, value="Weighted Total")
    header.font = _HEADER_FONT
    header.fill = _HEADER_FILL
    for i, b in enumerate(buckets, start=2):
        hist_ws.cell(row=i, column=start_col, value=b["count"])

    chart = BarChart()
    chart.type = "col"
    chart.style = 14
    chart.title = "Weighted total — class distribution"
    chart.y_axis.title = "Students"
    chart.x_axis.title = "Weighted score range (%)"
    chart.legend = None
    chart.data_labels = DataLabelList(showVal=True)
    data = Reference(hist_ws, min_col=start_col, max_col=start_col, min_row=1, max_row=1 + n)
    cats = Reference(hist_ws, min_col=1, min_row=2, max_row=1 + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    charts_ws.add_chart(chart, anchor)


def build_workbook(
    course_title: str,
    assignments: list[dict],
    students: list[dict],
    weightages: dict[str, float],
) -> BytesIO:
    """Build the workbook with Raw Data, Statistics, Charts, and a hidden _Histograms sheet."""
    wb = Workbook()
    wb.remove(wb.active)

    weighted_totals = compute_weighted_totals(students, weightages)

    raw_ws = wb.create_sheet("Raw Data")
    _write_raw_data(raw_ws, students, assignments, weighted_totals)

    stats_ws = wb.create_sheet("Statistics")
    _write_statistics(stats_ws, assignments, weightages)

    hist_ws = wb.create_sheet("_Histograms")
    hist_ws.sheet_state = "hidden"
    location = _write_histograms_sheet(hist_ws, assignments)

    charts_ws = wb.create_sheet("Charts")
    charts_ws.cell(row=1, column=1, value=f"Class performance — {course_title}").font = Font(
        bold=True, size=14
    )

    next_row = _add_histogram_charts(charts_ws, hist_ws, assignments, location)
    _add_submission_chart(charts_ws, stats_ws, len(assignments), f"B{next_row}")
    _add_trend_chart(charts_ws, stats_ws, len(assignments), f"B{next_row + 18}")

    weighted_col = 2 + len(assignments)  # next free column in _Histograms
    _add_weighted_total_chart(
        charts_ws, hist_ws, weighted_totals, weighted_col, f"B{next_row + 36}"
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
